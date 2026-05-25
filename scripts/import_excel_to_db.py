#!/usr/bin/env python3
"""
Импорт старых данных из Excel-выгрузки бота (/stats) в PostgreSQL или SQLite.

Ожидаются листы «сводка» и «ответы» (как при выгрузке stats_answers_*.xlsx).

Пример:
  set DATABASE_URL=postgresql://...
  pip install openpyxl psycopg2-binary
  python scripts/import_excel_to_db.py --excel C:\\Users\\User\\stats.xlsx

  python scripts/import_excel_to_db.py --excel stats.xlsx --dry-run
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sqlite3
import sys
from datetime import datetime, timezone

# --- сопоставление названий тестов из Excel → test_id в БД ---
TITLE_RULES: list[tuple[str, str]] = [
    ("ДДО", "klimov_self"),
    ("ОПГ", "opg"),
    ("Йовайши", "yovashi"),
    ("йоваши", "jovashi"),
    ("Таблица для ориентировочного", "jovashi"),
    ("ОПТ", "jovashi"),
    ("Кеттелл 16PF/C", "kettell_16pf_c"),
    ("Кеттелл 16PF", "kettell"),
    ("КОТ", "kot"),
    ("ЭН - 60", "en60"),
    ("ЭН - 57", "en57"),
    ("Голланд", "holland_riasec"),
    ("RIASEC", "holland_riasec"),
]

EXACT_TITLES = {
    "ДДО": "klimov_self",
    "ОПГ (опросник профессиональной готовности)": "opg",
    "Таблица для ориентировочного определения предпочтительности типа будущей профессии": "jovashi",
    "Йовайши (проф. склонности, модиф. Резапкиной)": "yovashi",
    "Кеттелл 16PF": "kettell",
    "Кеттелл 16PF/C (молодёжь)": "kettell_16pf_c",
    "КОТ (краткий ориентировочный тест)": "kot",
    "ЭН - 60": "en60",
    "ЭН - 57": "en57",
    "Голланд (RIASEC, пары профессий)": "holland_riasec",
}


def title_to_test_id(title: str) -> str:
    t = (title or "").strip()
    if t in EXACT_TITLES:
        return EXACT_TITLES[t]
    tl = t.lower()
    for needle, tid in TITLE_RULES:
        if needle.lower() in tl:
            return tid
    return "klimov_self"


def parse_vk_user_id(link: str) -> int | None:
    if not link:
        return None
    s = str(link).strip()
    m = re.search(r"vk\.com/id(\d+)", s, re.I)
    if m:
        return int(m.group(1))
    m = re.search(r"\bid(\d+)\b", s, re.I)
    if m:
        return int(m.group(1))
    return None


def parse_utc_ts(s) -> int:
    if s is None or s == "":
        return int(datetime.now(timezone.utc).timestamp())
    if isinstance(s, (int, float)):
        return int(s)
    txt = str(s).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(txt, fmt).replace(tzinfo=timezone.utc)
            return int(dt.timestamp())
        except ValueError:
            continue
    return int(datetime.now(timezone.utc).timestamp())


def parse_weights(raw) -> str:
    if raw is None or raw == "":
        return "{}"
    if isinstance(raw, dict):
        return json.dumps(raw, ensure_ascii=False)
    txt = str(raw).strip()
    try:
        obj = json.loads(txt)
        if isinstance(obj, dict):
            return json.dumps(obj, ensure_ascii=False)
    except json.JSONDecodeError:
        pass
    return "{}"


def _pg_url() -> str:
    url = (os.environ.get("DATABASE_URL") or "").strip()
    if not url:
        return ""
    if url.startswith("postgres://"):
        return "postgresql://" + url[len("postgres://") :]
    return url


def connect_db(sqlite_path: str | None):
    pg = _pg_url()
    if pg:
        import psycopg2

        return ("pg", psycopg2.connect(pg))
    if sqlite_path:
        return ("sqlite", sqlite3.connect(sqlite_path))
    print("Задайте DATABASE_URL или --sqlite", file=sys.stderr)
    sys.exit(1)


def load_sheet_rows(wb, names: tuple[str, ...]):
    for sn in wb.sheetnames:
        if sn.strip().lower() in names:
            return wb[sn]
    return None


def header_map(ws) -> dict[str, int]:
    row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    out = {}
    for i, v in enumerate(row):
        if v is None:
            continue
        out[str(v).strip()] = i
    return out


def col(h: dict[str, int], *aliases: str) -> int | None:
    for a in aliases:
        if a in h:
            return h[a]
    return None


def import_answers(ws, conn, backend: str, dry_run: bool) -> tuple[int, int, dict[tuple[int, str, int], int]]:
    h = header_map(ws)
    idx_link = col(h, "Ссылка на пользователя")
    idx_test = col(h, "Название теста")
    idx_sid = col(h, "session_id")
    idx_step = col(h, "step_index")
    idx_key = col(h, "answer_key")
    idx_lab = col(h, "answer_label")
    idx_q = col(h, "question_text")
    idx_w = col(h, "weights_json")
    idx_ts = col(h, "created_at_utc")
    if idx_link is None or idx_test is None:
        print("Лист «ответы»: не найдены колонки (Ссылка на пользователя, Название теста)", file=sys.stderr)
        return 0, 0, {}

    groups: dict[tuple[int, str, int], list[dict]] = {}
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(c is None for c in row):
            continue
        link = row[idx_link] if idx_link < len(row) else None
        uid = parse_vk_user_id(str(link or ""))
        if uid is None:
            skipped += 1
            continue
        title = row[idx_test] if idx_test is not None and idx_test < len(row) else ""
        tid = title_to_test_id(str(title or ""))
        try:
            old_sid = int(row[idx_sid]) if idx_sid is not None and row[idx_sid] is not None else 0
        except (TypeError, ValueError):
            old_sid = 0
        if old_sid == 0:
            old_sid = hash((uid, tid, row)) % 10_000_000
        step = int(row[idx_step] or 0) if idx_step is not None else 0
        key = str(row[idx_key] or "") if idx_key is not None else ""
        lab = str(row[idx_lab] or "") if idx_lab is not None else ""
        qtxt = str(row[idx_q] or "") if idx_q is not None else ""
        wjson = parse_weights(row[idx_w] if idx_w is not None else "{}")
        ts = parse_utc_ts(row[idx_ts] if idx_ts is not None else None)
        gkey = (uid, tid, old_sid)
        groups.setdefault(gkey, []).append(
            {
                "step_index": step,
                "answer_key": key,
                "answer_label": lab,
                "question_text": qtxt,
                "weights_json": wjson,
                "created_at": ts,
            }
        )

    sid_map: dict[tuple[int, str, int], int] = {}
    n_sessions = 0
    n_answers = 0
    cur = conn.cursor()

    for (uid, tid, old_sid), items in sorted(groups.items(), key=lambda x: min(i["created_at"] for i in x[1])):
        items.sort(key=lambda x: (x["created_at"], x["step_index"]))
        started = items[0]["created_at"]
        finished = items[-1]["created_at"]
        if dry_run:
            sid_map[(uid, tid, old_sid)] = -1
            n_sessions += 1
            n_answers += len(items)
            continue
        if backend == "pg":
            cur.execute(
                """
                INSERT INTO test_sessions (user_id, test_id, started_at, completed_at, status, final_scores_json)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
                """,
                (uid, tid, started, finished, "completed", "{}"),
            )
            new_sid = int(cur.fetchone()[0])
        else:
            cur.execute(
                """
                INSERT INTO test_sessions (user_id, test_id, started_at, completed_at, status, final_scores_json)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (uid, tid, started, finished, "completed", "{}"),
            )
            new_sid = int(cur.lastrowid)
        sid_map[(uid, tid, old_sid)] = new_sid
        n_sessions += 1
        for it in items:
            if backend == "pg":
                cur.execute(
                    """
                    INSERT INTO answer_log (
                        session_id, user_id, test_id, step_index, answer_key,
                        question_text, answer_label, weights_json, created_at
                    )
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    """,
                    (
                        new_sid,
                        uid,
                        tid,
                        it["step_index"],
                        it["answer_key"],
                        it["question_text"],
                        it["answer_label"],
                        it["weights_json"],
                        it["created_at"],
                    ),
                )
            else:
                cur.execute(
                    """
                    INSERT INTO answer_log (
                        session_id, user_id, test_id, step_index, answer_key,
                        question_text, answer_label, weights_json, created_at
                    )
                    VALUES (?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        new_sid,
                        uid,
                        tid,
                        it["step_index"],
                        it["answer_key"],
                        it["question_text"],
                        it["answer_label"],
                        it["weights_json"],
                        it["created_at"],
                    ),
                )
            n_answers += 1
    if not dry_run:
        conn.commit()
    print(f"  ответы: сессий {n_sessions}, записей answer_log {n_answers}, пропущено строк {skipped}")
    return n_sessions, n_answers, sid_map


def import_summary(ws, conn, backend: str, dry_run: bool) -> int:
    h = header_map(ws)
    idx_link = col(h, "Ссылка на пользователя")
    idx_test = col(h, "Название теста")
    idx_done = col(h, "Завершил")
    idx_dt = col(h, "Дата и время")
    idx_sum = col(h, "Итоги теста")
    if idx_link is None:
        print("Лист «сводка»: нет колонки «Ссылка на пользователя»", file=sys.stderr)
        return 0
    n = 0
    cur = conn.cursor()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        done = str(row[idx_done] or "").strip() if idx_done is not None else ""
        if not done.startswith("Да"):
            continue
        uid = parse_vk_user_id(str(row[idx_link] or ""))
        if uid is None:
            continue
        title = row[idx_test] if idx_test is not None else ""
        tid = title_to_test_id(str(title or ""))
        finished_at = parse_utc_ts(row[idx_dt] if idx_dt is not None else None)
        summary = str(row[idx_sum] or "") if idx_sum is not None else ""
        scores = {
            "_imported_from_excel": True,
            "summary_text": summary[:8000],
        }
        top3: list = []
        best = "imported"
        if dry_run:
            n += 1
            continue
        if backend == "pg":
            cur.execute(
                """
                INSERT INTO test_results (user_id, finished_at, scores_json, top3_json, best_type, test_id)
                VALUES (%s,%s,%s,%s,%s,%s)
                """,
                (
                    uid,
                    finished_at,
                    json.dumps(scores, ensure_ascii=False),
                    json.dumps(top3, ensure_ascii=False),
                    best,
                    tid,
                ),
            )
        else:
            cur.execute(
                """
                INSERT INTO test_results (user_id, finished_at, scores_json, top3_json, best_type, test_id)
                VALUES (?,?,?,?,?,?)
                """,
                (
                    uid,
                    finished_at,
                    json.dumps(scores, ensure_ascii=False),
                    json.dumps(top3, ensure_ascii=False),
                    best,
                    tid,
                ),
            )
        n += 1
    if not dry_run:
        conn.commit()
    print(f"  сводка (завершённые): test_results {n}")
    return n


def ensure_schema(conn, backend: str):
    if backend != "pg":
        return
    cur = conn.cursor()
    # минимальный DDL
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS test_results (
            id SERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            finished_at BIGINT NOT NULL,
            scores_json TEXT NOT NULL,
            top3_json TEXT NOT NULL,
            best_type TEXT NOT NULL DEFAULT '',
            test_id TEXT NOT NULL DEFAULT 'klimov_self'
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS test_sessions (
            id SERIAL PRIMARY KEY,
            user_id BIGINT NOT NULL,
            test_id TEXT NOT NULL,
            started_at BIGINT NOT NULL,
            completed_at BIGINT,
            status TEXT NOT NULL DEFAULT 'in_progress',
            final_scores_json TEXT
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS answer_log (
            id SERIAL PRIMARY KEY,
            session_id BIGINT NOT NULL,
            user_id BIGINT NOT NULL,
            test_id TEXT NOT NULL,
            step_index INTEGER NOT NULL,
            answer_key TEXT NOT NULL,
            question_text TEXT NOT NULL,
            answer_label TEXT NOT NULL,
            weights_json TEXT NOT NULL,
            created_at BIGINT NOT NULL
        )
        """
    )
    conn.commit()


def main():
    ap = argparse.ArgumentParser(description="Импорт Excel выгрузки бота в БД")
    ap.add_argument("--excel", required=True, help="Путь к .xlsx")
    ap.add_argument("--sqlite", default="", help="SQLite вместо DATABASE_URL")
    ap.add_argument("--dry-run", action="store_true", help="Только подсчёт, без записи")
    ap.add_argument(
        "--answers-only",
        action="store_true",
        help="Только лист «ответы» (пошаговые ответы)",
    )
    args = ap.parse_args()
    if not os.path.isfile(args.excel):
        print(f"Нет файла: {args.excel}", file=sys.stderr)
        sys.exit(1)
    try:
        import openpyxl
    except ImportError:
        print("pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)
    ws_ans = load_sheet_rows(wb, ("ответы", "answers"))
    ws_sum = load_sheet_rows(wb, ("сводка", "summary"))
    if ws_ans is None and ws_sum is None:
        print(f"Нет листов «ответы»/«сводка». Есть: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    backend, conn = connect_db(args.sqlite.strip() or None)
    ensure_schema(conn, backend)
    print(f"Режим: {backend}" + (" (dry-run)" if args.dry_run else ""))

    if ws_ans is not None:
        import_answers(ws_ans, conn, backend, args.dry_run)
    elif not args.answers_only:
        print("  лист «ответы» не найден — пошаговые ответы не импортируются")

    if ws_sum is not None and not args.answers_only:
        import_summary(ws_sum, conn, backend, args.dry_run)
    conn.close()
    print("Готово.")


if __name__ == "__main__":
    main()
